"""
=============================================================
  AI-Powered Attendance Report Agent
  KIET Group of Institutions — Dept. of CSE (AI / AIML)
  Reads the department's standard Excel attendance format
  Uses: OpenAI GPT-4o + openpyxl + pandas
=============================================================

Excel format expected (your department's standard sheet):
  Row 1  : Institution name
  Row 2  : Department name
  Row 3  : Class/Section/Session info
  Row 4  : Note
  Row 5  : Lecture numbers (1, 2, 3 ...)
  Row 6  : S.No | UNI ROLL NO | NAME/Date --> | date1 | date2 | ...
  Row 7+ : Student rows  (1 or 0 per lecture)
  Last   : "Total Students(Present)" row

Each date column = one lecture. Lecture dates appear as date objects
or strings in row 6 from column 4 onwards.

Usage:
  1. Add your OpenAI key to a .env file:  OPENAI_API_KEY=sk-xxxx
  2. Run:
     python attendance_agent.py --file attendance.xlsx
     python attendance_agent.py --file attendance.xlsx --threshold 80
     python attendance_agent.py --file attendance.xlsx --emails
"""

import argparse
import json
import os
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
from dotenv import load_dotenv
from openai import OpenAI

try:
    from fpdf import FPDF
except ModuleNotFoundError:
    FPDF = None

# Load API key from .env file
load_dotenv()

# ─── CONFIG ──────────────────────────────────────────────────────────────────

DEFAULT_THRESHOLD = 75
MODEL             = "gpt-4o"   # swap to "gpt-4o-mini" to cut cost ~15x


# ─── STEP 1: PARSE THE KIET EXCEL FORMAT ─────────────────────────────────────

def parse_excel(file_path: str, threshold: float) -> dict:
    """
    Reads the department's standard Excel sheet and returns a structured
    dict with all computed attendance data.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    # ── Extract metadata from header rows ──
    institution = str(all_rows[0][0]).strip() if all_rows[0][0] else "Institution"
    department  = str(all_rows[1][0]).strip() if all_rows[1][0] else "Department"
    class_info  = str(all_rows[2][0]).strip() if all_rows[2][0] else "Class Info"

    # ── Row 6 (index 5): header row with dates from column 4 (index 3) onward ──
    header_row = all_rows[5]

    # Collect lecture date labels (col index 3 onwards, skip None)
    def format_date(val):
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.strftime("%-d/%b")   # e.g. "2/May"
        return str(val).strip()

    lecture_dates = []
    lecture_col_indices = []
    for i, val in enumerate(header_row):
        if i < 3:
            continue
        label = format_date(val)
        if label:
            lecture_dates.append(label)
            lecture_col_indices.append(i)

    total_lectures = len(lecture_col_indices)

    # ── Rows 7+ : student data ──
    students = []
    for row in all_rows[6:]:
        # Stop at the "Total Students(Present)" footer row
        if row[2] and isinstance(row[2], str) and "Total" in row[2]:
            break
        # Skip fully empty rows
        if row[2] is None:
            continue
        name    = str(row[2]).strip()
        roll_no = str(row[1]).strip() if row[1] else ""

        # Collect attendance values for each lecture column
        attended = 0
        lecture_log = []
        for col_idx in lecture_col_indices:
            val = row[col_idx] if col_idx < len(row) else None
            present = 1 if val == 1.0 or val == 1 else 0
            attended += present
            lecture_log.append(present)

        # Only count lectures that actually had data (non-None in this student's row)
        conducted = sum(
            1 for col_idx in lecture_col_indices
            if col_idx < len(row) and row[col_idx] is not None
        )
        # Fallback: use total_lectures if no data gaps
        if conducted == 0:
            conducted = total_lectures

        pct = round((attended / conducted * 100), 1) if conducted > 0 else 0.0

        students.append({
            "name"        : name,
            "roll_no"     : roll_no,
            "attended"    : attended,
            "conducted"   : conducted,
            "pct"         : pct,
            "defaulter"   : pct < threshold,
            "lecture_log" : lecture_log,
        })

    # ── Aggregate stats ──
    total_students  = len(students)
    defaulters      = [s for s in students if s["defaulter"]]
    above_90        = [s for s in students if s["pct"] >= 90]
    avg_pct         = round(sum(s["pct"] for s in students) / total_students, 1) if students else 0

    # Per-lecture class attendance (how many present on each day)
    lecture_wise = []
    for i, date in enumerate(lecture_dates):
        present_count = sum(
            s["lecture_log"][i] for s in students if i < len(s["lecture_log"])
        )
        lecture_wise.append({
            "lecture_no" : i + 1,
            "date"       : date,
            "present"    : present_count,
            "absent"     : total_students - present_count,
            "pct"        : round(present_count / total_students * 100, 1) if total_students else 0,
        })

    return {
        "institution"    : institution,
        "department"     : department,
        "class_info"     : class_info,
        "threshold"      : threshold,
        "total_students" : total_students,
        "total_lectures" : total_lectures,
        "avg_pct"        : avg_pct,
        "students"       : students,
        "defaulters"     : defaulters,
        "above_90"       : above_90,
        "lecture_wise"   : lecture_wise,
        "lecture_dates"  : lecture_dates,
    }


# ─── STEP 2: GPT HELPERS ─────────────────────────────────────────────────────

def call_gpt(client: OpenAI, system: str, user: str) -> str:
    response = client.chat.completions.create(
        model=MODEL,
        messages=[
            {"role": "system", "content": system},
            {"role": "user",   "content": user},
        ],
        temperature=0.3,
        max_tokens=4096,
    )
    return response.choices[0].message.content.strip()


def validate_api_key(api_key: str):
    if not api_key.isascii():
        sys.exit(
            "[ERROR] OPENAI_API_KEY contains unsupported non-ASCII characters. "
            "Open .env and remove any accidental symbols such as '÷', then run again."
        )


# ─── STEP 3: REPORT GENERATORS ───────────────────────────────────────────────

def generate_summary_report(client: OpenAI, data: dict) -> str:
    system = (
        "You are a senior academic administrator AI producing an official attendance report "
        "for the HOD and faculty of a university CSE department in India (AKTU-affiliated). "
        "The report must be detailed, data-rich, and university-grade. "
        "Every section must contain actual numbers from the data — no vague statements. "
        "Use plain text with clear ALL-CAPS section headings. "
        "Use properly aligned ASCII tables for all tabular data. "
        "Do NOT write N/A if a category is empty — write a specific sentence using actual data instead. "
        "Do NOT treat dates with zero present+absent count as real lectures — those are unrecorded "
        "columns. Only analyse dates where at least one student has a recorded value."
    )

    # ── Pre-compute rich stats in Python ──
    pcts = [s["pct"] for s in data["students"]]
    sorted_pcts = sorted(pcts)
    median_pct = round(sorted_pcts[len(sorted_pcts) // 2], 1)

    buckets = {
        "90-100%"  : [s for s in data["students"] if s["pct"] >= 90],
        "75-89%"   : [s for s in data["students"] if 75 <= s["pct"] < 90],
        "60-74%"   : [s for s in data["students"] if 60 <= s["pct"] < 75],
        "Below 60%": [s for s in data["students"] if s["pct"] < 60],
    }

    # Valid lectures only (skip columns with no recorded data)
    valid_lectures = [
        lw for lw in data["lecture_wise"]
        if (lw["present"] + lw["absent"]) > 0
    ]
    worst_days = sorted(valid_lectures, key=lambda x: x["pct"])[:5]
    best_days  = sorted(valid_lectures, key=lambda x: -x["pct"])[:3]

    # Classes each defaulter needs to attend consecutively to cross threshold
    student_summary = []
    for s in data["students"]:
        classes_needed = 0
        if s["defaulter"]:
            t = data["threshold"] / 100
            val = (t * s["conducted"] - s["attended"]) / (1 - t)
            classes_needed = max(0, int(val) + 1)
        student_summary.append({
            "name"          : s["name"],
            "roll_no"       : s["roll_no"],
            "attended"      : s["attended"],
            "conducted"     : s["conducted"],
            "pct"           : s["pct"],
            "defaulter"     : s["defaulter"],
            "classes_needed": classes_needed,
        })

    user = f"""
Generate a detailed, university-grade attendance monitoring report for:

Institution : {data['institution']}
Department  : {data['department']}
Class       : {data['class_info']}
Threshold   : {data['threshold']}%

KEY STATISTICS (pre-computed — use these exact numbers):
  Total Students        : {data['total_students']}
  Valid Lectures Recorded: {len(valid_lectures)}
  Class Average          : {data['avg_pct']}%
  Median Attendance      : {median_pct}%
  Highest Attendance     : {max(pcts)}%
  Lowest Attendance      : {min(pcts)}%
  Total Defaulters       : {len(data['defaulters'])} ({round(len(data['defaulters'])/data['total_students']*100,1)}% of class)

DISTRIBUTION:
  Excellent (90-100%) : {len(buckets['90-100%'])} students ({round(len(buckets['90-100%'])/data['total_students']*100,1)}%)
  Good      (75-89%)  : {len(buckets['75-89%'])} students ({round(len(buckets['75-89%'])/data['total_students']*100,1)}%)
  At Risk   (60-74%)  : {len(buckets['60-74%'])} students ({round(len(buckets['60-74%'])/data['total_students']*100,1)}%)
  Critical (need immediate Attention)  (<60%)    : {len(buckets['Below 60%'])} students ({round(len(buckets['Below 60%'])/data['total_students']*100,1)}%)

VALID LECTURE-WISE DATA (only use these — ignore any with 0 present+absent):
{json.dumps(valid_lectures, indent=2)}

Worst 5 lecture days: {json.dumps(worst_days)}
Best 3 lecture days : {json.dumps(best_days)}

ALL STUDENT DATA (name, roll, attended, conducted, %, classes_needed to recover):
{json.dumps(student_summary, indent=2)}

=== GENERATE ALL 7 SECTIONS BELOW — do not skip any ===

1. EXECUTIVE SUMMARY
   - Class health status: Healthy / At Risk / Critical — with justification
   - All 6 key statistics listed above
   - Distribution across 4 attendance bands
   - Top 3 urgent concerns with specific numbers

2. ATTENDANCE DISTRIBUTION ANALYSIS
   Table: Band | Students | % of Class | Status
   Follow with 3-4 sentences of analysis — what does this distribution indicate?

3. DEFAULTER LIST (below {data['threshold']}%)
   Split into two sub-tables:
   A) CRITICAL — below 60%
      Table: Roll No | Name | Attended | Conducted | Attendance % | Classes Needed to Recover
   B) AT RISK — 60% to {data['threshold']-0.1}%
      Same table columns
   "Classes Needed to Recover" = consecutive future classes needed to cross {data['threshold']}%
   Use the classes_needed values from the data.

4. TOP PERFORMERS
   If no student is above 90%, do NOT write N/A.
   Instead write:
   "No student has crossed the 90% attendance mark in this monitoring period.
   The best attendance recorded is [X%] by [Name] ([Roll No]), followed by [Y%] by [Name]."
   List the top 5 students by attendance % with their actual numbers.

5. LECTURE-WISE ATTENDANCE TREND
   Table: Lec No | Date | Present | Absent | Class % (only valid lectures)
   - Best 3 days: specific numbers and dates
   - Worst 3 days: specific numbers and dates
   - Trend narrative: is attendance improving, declining, or inconsistent across the semester?
     Back this with actual numbers (e.g. first 5 lectures vs last 5 lectures average)

6. STUDENT-LEVEL INSIGHTS
   A) Students closest to threshold (need 5 or fewer classes to recover) — encouragement targets
   B) Students with severe risk (below 50%) — list with current % and classes needed
   C) Consistent attenders (above 95% if any, else above 85%) — recognise them

7. RECOMMENDATIONS FOR HOD
   Provide exactly 6 specific, actionable recommendations relevant to an AKTU-affiliated
   engineering college. Each must reference actual numbers from this data. Be precise —
   avoid generic advice. Examples of the required specificity:
   - "Issue formal detention warnings to the {len(buckets['Below 60%'])} students below 60% ..."
   - "The {len(buckets['60-74%'])} students in the 60-74% band need only X more classes ..."

Format: plain text, ALL-CAPS section headings, ASCII tables with aligned columns.
"""
    return call_gpt(client, system, user)


def generate_defaulter_emails(client: OpenAI, data: dict) -> list:
    emails = []
    system = (
        "You are drafting official attendance warning letters on behalf of a faculty member "
        f"at {data['institution']}, {data['department']}. "
        "Tone: formal, firm, respectful. Under 180 words per letter. "
        "Do not use placeholder brackets — use the actual data given."
    )
    for s in data["defaulters"]:
        user = f"""
Draft an official attendance warning letter for:

Student Name    : {s['name']}
University Roll : {s['roll_no']}
Classes Attended: {s['attended']} out of {s['conducted']} conducted
Attendance %    : {s['pct']}%
Minimum Required: {data['threshold']}%
Institution     : {data['institution']}
Department      : {data['department']}
Class           : {data['class_info']}

Include:
- Current attendance % and minimum required
- Shortfall in number of classes versus the tooper of the class
- Consequence: risk of detention from semester exams if not improved
- Request to meet the faculty advisor / class coordinator within 3 working days
- Polite but firm closing
"""
        body = call_gpt(client, system, user)
        emails.append({
            "name"    : s["name"],
            "roll_no" : s["roll_no"],
            "pct"     : s["pct"],
            "email"   : body,
        })
    return emails


def generate_lecture_insights(client: OpenAI, data: dict) -> str:
    system = (
        "You are an academic data analyst for a university CSE department in India. "
        "Produce a detailed, data-driven lecture-wise and student-wise attendance insights report. "
        "Every point must cite specific numbers. No generic observations. "
        "Ignore lecture dates where present + absent = 0 — those are unrecorded columns, not real absences."
    )

    valid_lectures = [
        lw for lw in data["lecture_wise"]
        if (lw["present"] + lw["absent"]) > 0
    ]

    pcts = [s["pct"] for s in data["students"]]
    chronic = [s for s in data["students"] if s["pct"] < 60]
    early_warn = [s for s in data["students"]
                  if data["threshold"] <= s["pct"] < data["threshold"] + 6]
    safe = [s for s in data["students"] if s["pct"] >= 85]

    # Week-over-week grouping (every 4 valid lectures = approximate week)
    chunks = [valid_lectures[i:i+4] for i in range(0, len(valid_lectures), 4)]
    chunk_avgs = [
        round(sum(l["pct"] for l in chunk) / len(chunk), 1)
        for chunk in chunks if chunk
    ]

    user = f"""
Produce a detailed attendance insights report for:
{data['class_info']} — {data['department']}

OVERVIEW:
  Total students        : {data['total_students']}
  Valid lectures        : {len(valid_lectures)} (dates with actual recorded data)
  Class average         : {data['avg_pct']}%
  Threshold             : {data['threshold']}%
  Chronic absentees (<60%) : {len(chronic)} students
  Early warning zone ({data['threshold']}-{data['threshold']+5}%) : {len(early_warn)} students
  Safe zone (>=85%)     : {len(safe)} students

VALID LECTURE-WISE DATA:
{json.dumps(valid_lectures, indent=2)}

PERIODIC AVERAGES (approx. groups of 4 lectures each):
{json.dumps(chunk_avgs)}

CHRONIC ABSENTEES (below 60%):
{json.dumps([{'name': s['name'], 'roll': s['roll_no'], 'pct': s['pct'], 'attended': s['attended'], 'conducted': s['conducted']} for s in chronic], indent=2)}

EARLY WARNING STUDENTS ({data['threshold']}% to {data['threshold']+5}%):
{json.dumps([{'name': s['name'], 'pct': s['pct']} for s in early_warn], indent=2)}

ALL STUDENT ATTENDANCE %:
{json.dumps([{'name': s['name'], 'pct': s['pct']} for s in sorted(data['students'], key=lambda x: x['pct'])], indent=2)}

=== GENERATE ALL 6 SECTIONS BELOW ===

1. SEMESTER ATTENDANCE TREND
   - Using the periodic averages above, describe whether attendance is improving,
     declining, or fluctuating across the semester
   - Cite the actual average for each period group
   - Identify the most productive period (highest average) and the worst period

2. LECTURE-WISE ANALYSIS
   - Full table: Lec No | Date | Present | Absent | Attendance %
   - Flag any lecture where class attendance dropped below 60% — these are red-flag days
   - Best 3 and worst 3 lectures with specific figures
   - Day-of-week or date patterns (e.g. end-of-month or post-holiday dips) if visible

3. CHRONIC ABSENTEE PROFILE
   - List all students below 60% with their numbers
   - Categorise: 40-59% (Severe), below 40% (Critical/Possible dropout risk)
   - For each: how many classes they have missed and how many they need to recover

4. EARLY WARNING RADAR
   - Students in the {data['threshold']}% to {data['threshold']+5}% band — at risk of becoming defaulters
   - For each: current %, classes attended, how many more they need to stay safe
   - These students need proactive nudging, not formal warnings yet

5. POSITIVE RECOGNITION
   - Students with highest attendance — list top 10 with their %
   - If any student attended every single recorded lecture, call that out specifically
   - Commentary: what % of the class maintained good standing (above {data['threshold']}%)?

6. ACTIONABLE INSIGHTS SUMMARY
   - 5 specific, numbered insights the class faculty can act on immediately
   - Each must cite a specific number or student group from this data
   - Focus on: intervention priority, recognition, scheduling observations

Format: plain text, ALL-CAPS section headings, ASCII tables with aligned columns.
"""
    return call_gpt(client, system, user)


# ─── STEP 4: SAVE OUTPUTS ────────────────────────────────────────────────────

def make_pdf_safe(text: str) -> str:
    """
    Convert generated text to plain ASCII so it works across both `fpdf`
    versions and the built-in fallback PDF writer.
    """
    replacements = {
        "—": "-",
        "–": "-",
        "÷": "/",
        "“": '"',
        "”": '"',
        "‘": "'",
        "’": "'",
        "✓": "OK",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")


def add_pdf_section(pdf: "FPDF", title: str, body: str):
    pdf.set_font("Arial", "B", 13)
    pdf.cell(0, 9, make_pdf_safe(title), ln=True)
    pdf.ln(1)
    pdf.set_font("Courier", size=8)
    pdf.multi_cell(0, 4, make_pdf_safe(body))
    pdf.ln(4)


def save_pdf_report(data: dict, summary: str, insights: str, emails: list,
                    output_dir: str, source_file: str) -> str:
    return save_basic_pdf_report(data, summary, insights, emails, output_dir, source_file)


def wrap_pdf_lines(text: str, width: int = 108) -> list:
    lines = []
    for raw_line in make_pdf_safe(text).splitlines():
        if not raw_line:
            lines.append("")
            continue
        while len(raw_line) > width:
            split_at = raw_line.rfind(" ", 0, width + 1)
            if split_at <= 0:
                split_at = width
            lines.append(raw_line[:split_at])
            raw_line = raw_line[split_at:].lstrip()
        lines.append(raw_line)
    return lines


def escape_pdf_text(text: str) -> str:
    return text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def is_report_heading(line: str) -> bool:
    stripped = line.strip()
    if not stripped or len(stripped) > 88:
        return False
    letters = [char for char in stripped if char.isalpha()]
    return bool(letters) and all(char.isupper() for char in letters)


def pdf_text(x: int, y: int, text: str, font: str = "F3", size: int = 8,
             color: tuple = (0.14, 0.18, 0.24)) -> str:
    r, g, b = color
    return (
        f"BT /{font} {size} Tf {r:.3f} {g:.3f} {b:.3f} rg "
        f"{x} {y} Td ({escape_pdf_text(text)}) Tj ET"
    )


def save_basic_pdf_report(data: dict, summary: str, insights: str, emails: list,
                          output_dir: str, source_file: str) -> str:
    report_sections = [
        ("ATTENDANCE SUMMARY REPORT", summary),
        ("LECTURE-WISE ATTENDANCE INSIGHTS", insights),
    ]
    if emails:
        letters = []
        for e in emails:
            letters.extend([
                f"TO   : {e['name']}",
                f"ROLL : {e['roll_no']}  |  Attendance: {e['pct']}%",
                "-" * 45,
                e["email"],
                "=" * 65,
                "",
            ])
        report_sections.append(("ATTENDANCE WARNING LETTERS", "\n".join(letters)))

    pages = []

    def new_page():
        page_no = len(pages) + 1
        ops = [
            "0.965 0.973 0.984 rg 0 0 595 842 re f",
            "0.067 0.176 0.337 rg 0 790 595 52 re f",
            "0.000 0.690 0.655 rg 36 784 523 4 re f",
            pdf_text(36, 811, "KIET ATTENDANCE REPORT", "F2", 9, (1, 1, 1)),
            pdf_text(503, 811, f"PAGE {page_no}", "F2", 9, (1, 1, 1)),
            pdf_text(36, 24, make_pdf_safe(data["department"]), "F1", 8, (0.34, 0.39, 0.47)),
        ]
        page = {"ops": ops, "y": 758}
        pages.append(page)
        return page

    def ensure_space(page, needed):
        if page["y"] - needed < 48:
            return new_page()
        return page

    def add_section_bar(page, title):
        page = ensure_space(page, 34)
        y = page["y"]
        page["ops"].append(f"0.067 0.176 0.337 rg 36 {y - 18} 523 24 re f")
        page["ops"].append(pdf_text(48, y - 11, make_pdf_safe(title), "F2", 11, (1, 1, 1)))
        page["y"] -= 38
        return page

    def add_body_line(page, line):
        if line.strip() and set(line.strip()) <= {"=", "-"}:
            return page
        if is_report_heading(line):
            page = ensure_space(page, 28)
            y = page["y"]
            page["ops"].append(f"0.890 0.925 0.965 rg 36 {y - 17} 523 22 re f")
            page["ops"].append(f"0.000 0.690 0.655 rg 36 {y - 17} 4 22 re f")
            page["ops"].append(pdf_text(48, y - 11, make_pdf_safe(line), "F2", 9, (0.067, 0.176, 0.337)))
            page["y"] -= 28
            return page

        page = ensure_space(page, 11)
        safe_line = make_pdf_safe(line)
        page["ops"].append(pdf_text(42, page["y"], safe_line, "F3", 7, (0.12, 0.15, 0.20)))
        page["y"] -= 10
        return page

    page = new_page()
    page["ops"].extend([
        "0.067 0.176 0.337 rg 36 650 523 88 re f",
        "0.000 0.690 0.655 rg 36 650 8 88 re f",
        pdf_text(58, 704, make_pdf_safe(data["institution"]), "F2", 18, (1, 1, 1)),
        pdf_text(58, 682, make_pdf_safe(data["department"]), "F1", 11, (0.88, 0.94, 0.98)),
        pdf_text(58, 664, make_pdf_safe(data["class_info"]), "F1", 9, (0.88, 0.94, 0.98)),
        pdf_text(42, 614, "REPORT SNAPSHOT", "F2", 10, (0.067, 0.176, 0.337)),
    ])
    snapshot = [
        ("Students", str(data["total_students"])),
        ("Lectures", str(data["total_lectures"])),
        ("Class Avg", f"{data['avg_pct']}%"),
        ("Defaulters", str(len(data["defaulters"]))),
    ]
    card_x = [42, 171, 300, 429]
    for x, (label, value) in zip(card_x, snapshot):
        page["ops"].append(f"1 1 1 rg {x} 548 112 52 re f")
        page["ops"].append(f"0.827 0.863 0.906 RG {x} 548 112 52 re S")
        page["ops"].append(pdf_text(x + 10, 580, label.upper(), "F2", 8, (0.34, 0.39, 0.47)))
        page["ops"].append(pdf_text(x + 10, 558, value, "F2", 18, (0.067, 0.176, 0.337)))
    page["y"] = 510

    for section_title, section_body in report_sections:
        page = add_section_bar(page, section_title)
        for line in wrap_pdf_lines(section_body, width=108):
            page = add_body_line(page, line)

    objects = []
    page_refs = []
    first_font_obj_no = 3 + (len(pages) * 2)
    font_refs = {
        "F1": first_font_obj_no,
        "F2": first_font_obj_no + 1,
        "F3": first_font_obj_no + 2,
    }

    objects.append("<< /Type /Catalog /Pages 2 0 R >>")
    kids = []
    for index, page in enumerate(pages):
        page_obj_no = 3 + (index * 2)
        content_obj_no = page_obj_no + 1
        page_refs.append((page_obj_no, content_obj_no, page))
        kids.append(f"{page_obj_no} 0 R")
    objects.append(f"<< /Type /Pages /Kids [{' '.join(kids)}] /Count {len(pages)} >>")

    for page_obj_no, content_obj_no, page in page_refs:
        objects.append(
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
            f"/Contents {content_obj_no} 0 R /Resources << /Font << "
            f"/F1 {font_refs['F1']} 0 R /F2 {font_refs['F2']} 0 R /F3 {font_refs['F3']} 0 R "
            f">> >> >>"
        )
        stream = "\n".join(page["ops"])
        objects.append(f"<< /Length {len(stream.encode('latin-1'))} >>\nstream\n{stream}\nendstream")

    objects.append("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    objects.append("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>")
    objects.append("<< /Type /Font /Subtype /Type1 /BaseFont /Courier >>")

    pdf_parts = ["%PDF-1.4\n"]
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(sum(len(part.encode("latin-1")) for part in pdf_parts))
        pdf_parts.append(f"{index} 0 obj\n{obj}\nendobj\n")

    xref_offset = sum(len(part.encode("latin-1")) for part in pdf_parts)
    pdf_parts.append(f"xref\n0 {len(objects) + 1}\n")
    pdf_parts.append("0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf_parts.append(f"{offset:010d} 00000 n \n")
    pdf_parts.append(
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\n"
        f"startxref\n{xref_offset}\n%%EOF"
    )

    pdf_name = f"{Path(source_file).stem}.pdf"
    pdf_path = os.path.join(output_dir, pdf_name)
    with open(pdf_path, "wb") as f:
        f.write("".join(pdf_parts).encode("latin-1"))
    return pdf_name

def save_outputs(data: dict, summary: str, insights: str,
                 emails: list, output_dir: str, source_file: str):
    os.makedirs(output_dir, exist_ok=True)

    # 1. Summary report
    with open(f"{output_dir}/attendance_report.txt", "w", encoding="utf-8") as f:
        f.write(f"{data['institution']}\n")
        f.write(f"{data['department']}\n")
        f.write(f"{data['class_info']}\n")
        f.write("=" * 65 + "\n")
        f.write("ATTENDANCE SUMMARY REPORT\n")
        f.write("=" * 65 + "\n\n")
        f.write(summary)

    # 2. Insights
    with open(f"{output_dir}/attendance_insights.txt", "w", encoding="utf-8") as f:
        f.write(f"{data['institution']} — {data['department']}\n")
        f.write("LECTURE-WISE ATTENDANCE INSIGHTS\n")
        f.write("=" * 65 + "\n\n")
        f.write(insights)

    # 3. Warning emails
    if emails:
        with open(f"{output_dir}/warning_letters.txt", "w", encoding="utf-8") as f:
            f.write("ATTENDANCE WARNING LETTERS\n")
            f.write(f"Generated for {len(emails)} defaulters\n")
            f.write("=" * 65 + "\n\n")
            for e in emails:
                f.write(f"TO   : {e['name']}\n")
                f.write(f"ROLL : {e['roll_no']}  |  Attendance: {e['pct']}%\n")
                f.write("-" * 45 + "\n")
                f.write(e["email"])
                f.write("\n\n" + "=" * 65 + "\n\n")

    # 4. Defaulter CSV (useful for HOD/ERP import)
    if data["defaulters"]:
        rows = [
            {
                "Roll No"           : s["roll_no"],
                "Name"              : s["name"],
                "Classes Attended"  : s["attended"],
                "Classes Conducted" : s["conducted"],
                "Attendance %"      : s["pct"],
                "Status"            : "DEFAULTER",
            }
            for s in sorted(data["defaulters"], key=lambda x: x["pct"])
        ]
        pd.DataFrame(rows).to_csv(f"{output_dir}/defaulters.csv", index=False)

    pdf_name = save_pdf_report(data, summary, insights, emails, output_dir, source_file)

    print(f"\n[✓] Output saved to: {output_dir}/")
    print(f"    attendance_report.txt    — Full summary report for HOD")
    print(f"    attendance_insights.txt  — Lecture-wise trend analysis")
    print(f"    {pdf_name:<24} — Combined PDF report")
    if emails:
        print(f"    warning_letters.txt      — {len(emails)} warning letters")
    if data["defaulters"]:
        print(f"    defaulters.csv           — Defaulter list (HOD/ERP ready)")


# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="AI Attendance Report Agent — KIET CSE Dept Excel Format"
    )
    parser.add_argument("--file",      required=True,
                        help="Path to the department Excel attendance sheet (.xlsx)")
    parser.add_argument("--threshold", type=float, default=DEFAULT_THRESHOLD,
                        help=f"Minimum attendance %% (default: {DEFAULT_THRESHOLD})")
    parser.add_argument("--emails",    action="store_true",
                        help="Generate warning letters for each defaulter")
    parser.add_argument("--output",    default="attendance_output",
                        help="Output folder (default: attendance_output/)")
    args = parser.parse_args()

    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        sys.exit("[ERROR] OPENAI_API_KEY not found. Add it to your .env file:\n"
                 "        OPENAI_API_KEY=sk-xxxx")
    validate_api_key(api_key)
    client = OpenAI(api_key=api_key)

    # ── 1. Parse Excel ──
    print(f"\n[1/4] Reading Excel sheet: {args.file}")
    data = parse_excel(args.file, args.threshold)
    print(f"      Institution : {data['institution']}")
    print(f"      Department  : {data['department']}")
    print(f"      Class       : {data['class_info']}")
    print(f"      Students    : {data['total_students']}")
    print(f"      Lectures    : {data['total_lectures']}")
    print(f"      Defaulters  : {len(data['defaulters'])} (below {args.threshold}%)")
    print(f"      Class avg   : {data['avg_pct']}%")

    # ── 2. Summary report ──
    print("\n[2/4] Generating attendance summary report...")
    summary = generate_summary_report(client, data)

    # ── 3. Lecture insights ──
    print("[3/4] Generating lecture-wise insights...")
    insights = generate_lecture_insights(client, data)

    # ── 4. Warning emails ──
    emails = []
    if args.emails:
        if data["defaulters"]:
            print(f"[4/4] Drafting warning letters for {len(data['defaulters'])} defaulters...")
            emails = generate_defaulter_emails(client, data)
        else:
            print("[4/4] No defaulters found — skipping warning letters.")
    else:
        print("[4/4] Skipping warning letters (use --emails to enable).")

    # ── 5. Save ──
    save_outputs(data, summary, insights, emails, args.output, args.file)

    # ── Console quick view ──
    print("\n" + "─" * 65)
    print(f"DEFAULTERS ({len(data['defaulters'])} students below {args.threshold}%)")
    print("─" * 65)
    if data["defaulters"]:
        print(f"  {'Roll No':<20} {'Name':<25} {'Attended':>8} {'Total':>6} {'%':>6}")
        print("  " + "-" * 63)
        for s in sorted(data["defaulters"], key=lambda x: x["pct"]):
            print(f"  {s['roll_no']:<20} {s['name']:<25} {s['attended']:>8} {s['conducted']:>6} {s['pct']:>5}%")
    else:
        print("  No defaulters found. ")
    print("─" * 65)


if __name__ == "__main__":
    main()
